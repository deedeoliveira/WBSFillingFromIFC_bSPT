import os
import threading
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from ifcopenshell.util.element import get_psets

from app.gui.wbs_helpers import (
    find_wbs_columns,
    unpack_core_columns,
    split_levels,
    detect_relevant_leaves,
)
from app.gui.views.home import HomePage
from app.gui.views.wbs_editor import WBSPage
from app.gui.views.qty import QtyPage
from app.gui.views.report import ReportPage
from app.core.structural_engine import IFCInvestigator
from app.core.structural_engine import (
    load_and_migrate_rules,
    migrate_rule_v1_to_v2,
    NO_ELEMENTS_FOUND,
)


class WBSApp(tk.Tk):

    def __init__(self):
        super().__init__()

        try:
            from app import __version__
            version_str = f" | v{__version__}"
        except ImportError:
            version_str = ""

        self.title(f"Extração de Quantidades IFC → WBS{version_str}")

        self.minsize(1200, 800)
        try:
            self.state("zoomed")
        except Exception:
            pass
        self._fullscreen = False
        self.bind("<F11>", self._toggle_fullscreen)
        self.bind("<Escape>", self._exit_fullscreen)

        self.df_raw: pd.DataFrame | None = None
        self.df_desc0: pd.Series | None = None
        self.wbs_cols = {}
        self.col_nivel: str | None = "Nível"
        self.col_wbs: str | None = "WBS"
        self.col_desc: str | None = "DESCRIÇÃO"

        self.wbs_xlsx_var = tk.StringVar(value="")
        self.ifc_var = tk.StringVar(value="")
        self.out_var = tk.StringVar(value="")
        self.map_var = tk.StringVar(value="")

        self.rules = {}
        self.wbs_finalized = False

        self.ifc_file = None
        self.ifc_path_loaded = None
        self.inv = IFCInvestigator()

        self._previous_tab_index = None

        self._build_ui()

    def pick_file(self, var: tk.StringVar, title: str, patterns):
        path = filedialog.askopenfilename(title=title, filetypes=patterns)
        if path:
            var.set(path)

    def pick_dir(self, var: tk.StringVar, title: str):
        path = filedialog.askdirectory(title=title)
        if path:
            var.set(path)

    def _build_ui(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=8, pady=8)

        try:
            self.page_home = HomePage(self.notebook, self)
        except Exception as e:
            raise

        try:
            self.page_wbs = WBSPage(self.notebook, self)
        except Exception as e:
            raise

        try:
            self.page_qty = QtyPage(self.notebook, self)
        except Exception as e:
            raise

        try:
            self.page_report = ReportPage(self.notebook, self)
        except Exception as e:
            self.page_report = None

        try:
            self.notebook.add(self.page_home, text="Home")
        except Exception:
            pass

        try:
            self.notebook.add(self.page_wbs, text="WBS e descrição")
        except Exception:
            pass

        try:
            self.notebook.add(self.page_qty, text="Mapeamento IFC")
        except Exception:
            pass

        if self.page_report is not None:
            try:
                self.notebook.add(self.page_report, text="Extrair quantidades e Gerar WBS preenchido")
            except Exception:
                pass

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)

    def _get_tab_index(self, page_widget):
        try:
            for i in range(self.notebook.index('end')):
                if self.notebook.nametowidget(self.notebook.tabs()[i]) == page_widget:
                    return i
            return None
        except Exception:
            return None

    def go_home(self):
        try:
            idx = self._get_tab_index(self.page_home)
            if idx is not None:
                self.notebook.select(idx)
        except Exception:
            pass

    def go_wbs(self):
        try:
            idx = self._get_tab_index(self.page_wbs)
            if idx is not None:
                self.notebook.select(idx)
        except Exception:
            pass

    def open_mapping(self, source: str = "home"):
        try:
            if hasattr(self.page_qty, "set_mode"):
                self.page_qty.set_mode(source)
            idx = self._get_tab_index(self.page_qty)
            if idx is not None:
                self.notebook.select(idx)
        except Exception:
            pass

    def open_extract(self, source: str = "home"):
        try:
            if hasattr(self.page_report, "set_mode"):
                self.page_report.set_mode(source)
            idx = self._get_tab_index(self.page_report)
            if idx is not None:
                self.notebook.select(idx)
        except Exception:
            pass

    def go_mapping(self):
        self.open_mapping("home")

    def go_extract(self):
        self.open_extract("home")

    def has_user_descriptions(self) -> bool:
        if self.df_raw is None or self.col_nivel is None:
            return False
        try:
            pairs = detect_relevant_leaves(
                self.df_raw, self.col_nivel, self.col_wbs, self.col_desc, self.df_desc0
            )
            return bool(pairs)
        except Exception:
            return False

    def has_ifc_mapping(self) -> bool:
        return any(self.rules.values())

    def ensure_ifc_loaded(self, path: str):
        if not path:
            raise ValueError("Caminho IFC vazio.")
        if self.ifc_path_loaded == path and self.ifc_file is not None:
            return self.ifc_file
        self.ifc_file = self.inv.open_ifc(path)
        self.ifc_path_loaded = path
        return self.ifc_file

    def _toggle_fullscreen(self, event=None):
        self._fullscreen = not getattr(self, "_fullscreen", False)
        self.attributes("-fullscreen", self._fullscreen)

    def _exit_fullscreen(self, event=None):
        self._fullscreen = False
        self.attributes("-fullscreen", False)

    def _clear_memory_for_extraction(self):
        self.df_raw = None
        self.df_desc0 = None
        self.ifc_file = None
        self.ifc_path_loaded = None

    def _on_tab_changed(self, event):
        try:
            current_tab = self.notebook.select()
            current_tab_index = self.notebook.index(current_tab)

            if current_tab_index == 3 and self._previous_tab_index in [1, 2]:
                self._clear_memory_for_extraction()

            if current_tab_index == 0:
                if hasattr(self.page_home, 'refresh_on_show'):
                    self.page_home.refresh_on_show()

            self._previous_tab_index = current_tab_index

        except Exception:
            pass

    def _collect_grouping(self, details: list, g_pset: str, g_prop: str) -> tuple[list, dict]:
        seen = {}
        grp_sums = {}

        for det in details:
            e = det.get("element")
            val = det.get("valor", 0.0)
            if e is None:
                continue
            try:
                psets = get_psets(e) or {}
                gval = psets.get(g_pset, {}).get(g_prop)
                if gval is None:
                    continue
                sval = str(gval).strip()
                if not sval:
                    continue
                if sval not in seen:
                    seen[sval] = len(seen)
                grp_sums[sval] = grp_sums.get(sval, 0.0) + float(val or 0.0)
            except Exception:
                pass

        grp_vals = sorted(seen.keys(), key=lambda v: seen[v])
        return grp_vals, grp_sums

    def _get_leaf_desc_and_unit(self, code: str):
        from app.gui.wbs_helpers import split_levels
        try:
            df = self.df_raw.copy()
            niv = split_levels(df, self.col_nivel)
            parent_rows = df[
                (df[self.col_wbs].astype(str).str.strip() == code) &
                (niv < 10)
            ]
            if parent_rows.empty:
                return "", "n/a"
            leaf_idx = parent_rows.index[0]
            next_idx = leaf_idx + 1
            if next_idx < len(df) and niv.iat[next_idx] == 10:
                desc = str(df.at[next_idx, self.col_desc]) if self.col_desc in df.columns else ""
                unit = ""
                unit_candidates = ["UNIDADES", "UNID.", "UNID", "UNIT"]
                wbs_cols_map = getattr(self, "wbs_cols", {})
                col_u = wbs_cols_map.get("col_unidades")
                if col_u and col_u in df.columns:
                    unit_candidates = [col_u] + unit_candidates
                for uc in unit_candidates:
                    if uc in df.columns:
                        val = df.at[next_idx, uc]
                        if pd.notna(val) and str(val).strip() not in ("", "n/a", "nan"):
                            unit = str(val).strip()
                            break
                return desc, unit
        except Exception:
            pass
        return "", "n/a"

    def run_generate_report(self, log_widget, on_finish=None):

        def log(msg):
            log_widget.configure(state="normal")
            log_widget.insert("end", msg + ("\n" if not msg.endswith("\n") else ""))
            log_widget.see("end")
            log_widget.configure(state="disabled")

        def worker():
            import json

            from app.gui.wbs_helpers import split_levels, find_wbs_columns, unpack_core_columns
            from app.core.structural_engine import (
                load_and_migrate_rules, migrate_rule_v1_to_v2,
                NO_ELEMENTS_FOUND,
            )

            try:
                if not self.rules:
                    raise RuntimeError("Carregue um mapeamento (JSON).")

                self.rules = {c: migrate_rule_v1_to_v2(r) for c, r in self.rules.items()}

                wbs_path = self.wbs_xlsx_var.get().strip()
                if not wbs_path or not Path(wbs_path).is_file():
                    raise RuntimeError("Selecione um ficheiro WBS válido.")

                if getattr(self, "wbs_finalized", False) and self.df_raw is not None:
                    log("Usando WBS editado (com descrições do utilizador)")
                    if not all([self.col_wbs, self.col_desc, self.col_nivel]):
                        raise RuntimeError("Colunas WBS não detectadas. Recarregue o WBS.")
                else:
                    log(f"Carregando WBS: {Path(wbs_path).name}")
                    self.df_raw = pd.read_excel(wbs_path, header=1)
                    cols = find_wbs_columns(self.df_raw)
                    self.col_wbs, self.col_desc, self.col_nivel = unpack_core_columns(cols)
                    self.wbs_cols = cols
                    if not all([self.col_wbs, self.col_desc, self.col_nivel]):
                        raise RuntimeError("Não foi possível detectar as colunas WBS/Descrição/Nível.")
                    self.df_desc0 = self.df_raw[self.col_desc].copy()

                ifc_path = self.ifc_var.get().strip()
                if not ifc_path or not Path(ifc_path).is_file():
                    raise RuntimeError("Selecione um ficheiro IFC válido.")
                log(f"Carregando IFC: {Path(ifc_path).name}")
                self.ifc_file = None
                self.ifc_path_loaded = None
                self.inv.open_ifc(ifc_path)

                out_dir = Path(self.out_var.get().strip() or Path.home())
                out_dir.mkdir(parents=True, exist_ok=True)

                project_info = self.inv.get_project_info()
                ifc_project  = project_info.get("project",  "n/a")
                ifc_site     = project_info.get("site",     "n/a")
                ifc_building = project_info.get("building", "n/a")

                def _sort_key(code):
                    return tuple(int(x) if x.isdigit() else x for x in code.split("."))

                code_to_qty:       dict[str, float] = {}
                code_to_groupvals: dict[str, list]  = {}
                code_to_groupqtys: dict[str, dict]  = {}
                code_to_desc_idx:  dict[str, int]   = {}
                code_to_unit:      dict[str, str]   = {}
                no_elements_codes: list[str]         = []
                all_details:       list[dict]        = []

                log("\nExtraindo quantidades por código WBS:")

                for code in sorted(self.rules.keys(), key=_sort_key):
                    rule = self.rules[code]
                    qty_type = rule.get("quantity", {}).get("type", "prop")

                    try:
                        total, details, found_any = self.inv.extract_quantities(rule)

                        if not found_any:
                            no_elements_codes.append(code)
                            code_to_qty[code] = 0.0
                            log(f" - {code}: [ELEMENTOS NÃO ENCONTRADOS]")
                            _niv_s = split_levels(self.df_raw, self.col_nivel)
                            _parent_rows = self.df_raw[
                                (self.df_raw[self.col_wbs].astype(str).str.strip() == code) &
                                (_niv_s < 10)
                            ]
                            if not _parent_rows.empty:
                                _leaf_idx = _parent_rows.index[0]
                                _next_idx = _leaf_idx + 1
                                if _next_idx < len(self.df_raw) and _niv_s.iat[_next_idx] == 10:
                                    code_to_desc_idx[code] = _next_idx
                                    _, _unit = self._get_leaf_desc_and_unit(code)
                                    code_to_unit[code] = _unit
                            continue

                        code_to_qty[code] = float(total)
                        log(f" - {code}: {total}" + (" (contagem)" if qty_type == "count" else ""))

                        agr    = rule.get("agrupamento") or {}
                        g_pset = agr.get("pset", "")
                        g_prop = agr.get("prop", "")
                        if g_pset and g_prop:
                            grp_vals, grp_sums = self._collect_grouping(details, g_pset, g_prop)
                        else:
                            grp_vals, grp_sums = [], {}

                        code_to_groupvals[code] = grp_vals
                        code_to_groupqtys[code] = grp_sums

                        desc, unit = self._get_leaf_desc_and_unit(code)

                        niv_s = split_levels(self.df_raw, self.col_nivel)
                        parent_rows = self.df_raw[
                            (self.df_raw[self.col_wbs].astype(str).str.strip() == code) &
                            (niv_s < 10)
                        ]
                        if not parent_rows.empty:
                            leaf_idx = parent_rows.index[0]
                            next_idx = leaf_idx + 1
                            if next_idx < len(self.df_raw) and niv_s.iat[next_idx] == 10:
                                code_to_desc_idx[code] = next_idx
                                code_to_unit[code]     = unit

                        for det in details:
                            e    = det.get("element")
                            guid = det.get("guid", "n/a")
                            val  = det.get("valor")
                            mat  = self.inv.get_element_material(e) if e else "n/a"
                            cls_code  = self.inv.get_classification_code(e) if e else "n/a"
                            storey    = self.inv.get_building_storey(e) if e else "n/a"

                            ifc_cls = "n/a"; predef = "n/a"; objtype = "n/a"
                            if e is not None:
                                ifc_cls = e.is_a()
                                predef  = str(getattr(e, "PredefinedType", "n/a"))
                                objtype = str(getattr(e, "ObjectType", "n/a") or "n/a")

                            group_value = None
                            if g_pset and g_prop and e is not None:
                                try:
                                    psets = get_psets(e) or {}
                                    gv    = psets.get(g_pset, {}).get(g_prop)
                                    if gv is not None:
                                        sv = str(gv).strip()
                                        group_value = sv if sv else None
                                except Exception:
                                    pass

                            all_details.append({
                                "ifc_project":    ifc_project,
                                "ifc_site":       ifc_site,
                                "ifc_building":   ifc_building,
                                "wbs_codigo":     code,
                                "parent_code":    code,
                                "group_value":    group_value,
                                "descricao":      desc,
                                "ifc_class":      ifc_cls,
                                "predefinedtype": predef,
                                "objecttype":     objtype,
                                "wbs_grouping":   "n/a",
                                "material":       mat,
                                "classification_code": cls_code,
                                "buildingstorey": storey,
                                "ifc_guid":       guid,
                                "ifc_valor":      val,
                                "unidade":        unit,
                                "qty_type":       qty_type,
                            })

                    except Exception as e:
                        log(f" - {code}: erro ({e})")

                if no_elements_codes:
                    log(f"\n⚠  Sem elementos encontrados para {len(no_elements_codes)} código(s):")
                    for c in no_elements_codes:
                        log(f"    • {c}")

                df      = self.df_raw.copy()
                lvl     = split_levels(df, self.col_nivel)
                col_qty = "QDTE."
                col_uni = "UNID."

                for c in (col_qty, col_uni):
                    if c not in df.columns:
                        df[c] = ""

                wbs_cols_map = getattr(self, "wbs_cols", {})
                col_unidades = wbs_cols_map.get("col_unidades", col_uni)
                if col_unidades and col_unidades in df.columns and col_unidades != col_uni:
                    df[col_uni] = df[col_unidades]

                keep_cols = [self.col_wbs, self.col_desc, col_qty, col_uni]
                df_export = df[[c for c in keep_cols if c in df.columns]].copy()
                if col_uni not in df_export.columns and col_unidades in df_export.columns:
                    df_export = df_export.rename(columns={col_unidades: col_uni})

                idx10_to_parent = {v: k for k, v in code_to_desc_idx.items()}
                self.last_code_extensions = {}

                col_qnt = "QUANTIDADE"

                def _build_rows(include_not_found: bool):
                    rows  = []
                    kinds = []
                    ancestor_written = set()

                    codes_to_include = set()
                    for code in self.rules:
                        if not include_not_found and code in no_elements_codes:
                            continue
                        codes_to_include.add(code)
                        parts = code.split(".")
                        for k in range(1, len(parts)):
                            codes_to_include.add(".".join(parts[:k]))

                    for idx, row in df_export.iterrows():
                        lv = lvl.iat[idx] if idx < len(lvl) else None
                        is_desc10 = (lv == 10)
                        wbs_val   = str(row.get(self.col_wbs, "") or "").strip()

                        if not is_desc10:
                            if wbs_val not in codes_to_include:
                                continue
                            new_row = row.copy()
                            new_row[col_qty] = ""
                            rows.append(new_row)
                            kinds.append("wbs")
                            continue

                        parent_code = idx10_to_parent.get(idx)
                        if not parent_code:
                            continue
                        if not include_not_found and parent_code in no_elements_codes:
                            continue
                        if parent_code not in codes_to_include:
                            continue

                        new_row = row.copy()
                        desc_ext = f"{parent_code}.01"
                        new_row[self.col_wbs] = desc_ext

                        is_no_elem = parent_code in no_elements_codes
                        group_vals = code_to_groupvals.get(parent_code) or []
                        groups_map = {
                            str(gval): f"{desc_ext}.{k:02d}"
                            for k, gval in enumerate(group_vals, start=1)
                        }
                        self.last_code_extensions[parent_code] = {
                            "desc": desc_ext, "groups": groups_map
                        }

                        if is_no_elem:
                            new_row[col_qty] = ""
                            if col_qnt in new_row.index:
                                new_row[col_qnt] = "[ELEMENTOS NÃO ENCONTRADOS]"
                            new_row[col_uni] = ""
                        else:
                            total_qty = code_to_qty.get(parent_code, "")
                            new_row[col_qty] = "" if group_vals else total_qty
                            if col_qnt in new_row.index:
                                new_row[col_qnt] = "" if group_vals else total_qty
                            new_row[col_uni] = code_to_unit.get(parent_code, "")

                        rows.append(new_row)
                        kinds.append("desc10")

                        for gval in group_vals:
                            ins = {c: "" for c in df_export.columns}
                            ins[self.col_wbs] = groups_map.get(str(gval), "")
                            ins[self.col_desc] = str(gval)
                            ins[col_qty] = code_to_groupqtys.get(parent_code, {}).get(str(gval), 0.0)
                            ins[col_qnt] = ins[col_qty]
                            ins[col_uni] = code_to_unit.get(parent_code, "")
                            rows.append(pd.Series(ins, index=df_export.columns))
                            kinds.append("insert")

                    if include_not_found:
                        handled = {idx10_to_parent.get(i) for i in range(len(df_export))}
                        for code in no_elements_codes:
                            if code not in handled:
                                parts = code.split(".")
                                for k in range(1, len(parts) + 1):
                                    anc = ".".join(parts[:k])
                                    if anc not in codes_to_include:
                                        continue
                                    if any(str(r.get(self.col_wbs, "")).strip() == anc
                                           for r in rows if hasattr(r, "get")):
                                        continue
                                    anc_rows = df_export[
                                        df_export[self.col_wbs].astype(str).str.strip() == anc
                                    ]
                                    if not anc_rows.empty:
                                        nr = anc_rows.iloc[0].copy()
                                        nr[col_qty] = ""
                                        rows.append(nr)
                                        kinds.append("wbs")
                                desc_ext = f"{code}.01"
                                ins = {c: "" for c in df_export.columns}
                                ins[self.col_wbs]  = desc_ext
                                ins[self.col_desc] = ""
                                ins[col_qnt]       = "[ELEMENTOS NÃO ENCONTRADOS]"
                                ins[col_uni]       = ""
                                rows.append(pd.Series(ins, index=df_export.columns))
                                kinds.append("desc10")
                                self.last_code_extensions[code] = {
                                    "desc": desc_ext, "groups": {}
                                }

                    return rows, kinds

                df_export[col_qnt] = ""
                rows_full, kinds_full   = _build_rows(include_not_found=True)
                rows_found, kinds_found = _build_rows(include_not_found=False)

                cols_full  = list(df_export.columns)
                cols_found = [c for c in cols_full if c != col_qnt]

                df_full  = pd.DataFrame(rows_full,  columns=cols_full)
                df_found = pd.DataFrame(rows_found, columns=cols_found)

                def _style_sheet(ws, row_kinds):
                    header_fill  = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    header_font  = Font(bold=True, color="FFFFFF", size=12)
                    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    gray_fill    = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    white_fill   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                    warn_fill    = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                    border       = Border(
                        left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"),  bottom=Side(style="thin"),
                    )
                    data_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    for cell in ws[1]:
                        if cell.value:
                            cell.fill = header_fill; cell.font = header_font
                            cell.alignment = header_align; cell.border = border
                    for i, kind in enumerate(row_kinds, start=2):
                        is_warn = any(
                            "NÃO ENCONTRADOS" in str(ws.cell(row=i, column=c).value or "")
                            for c in range(1, ws.max_column + 1)
                        )
                        fill = warn_fill if is_warn else (gray_fill if kind == "wbs" else white_fill)
                        for cell in ws[i]:
                            cell.alignment = data_align; cell.border = border; cell.fill = fill
                    ws.column_dimensions["A"].width = 20
                    ws.column_dimensions["B"].width = 50
                    ws.column_dimensions["C"].width = 15
                    ws.column_dimensions["D"].width = 12

                ifc_stem = Path(ifc_path).stem if ifc_path else "output"

                out_path_found = out_dir / f"MapaQuantidadesTrabalhos_{ifc_stem}.xlsx"
                n = 1
                while out_path_found.exists():
                    out_path_found = out_dir / f"MapaQuantidadesTrabalhos_{ifc_stem}({n}).xlsx"
                    n += 1
                with pd.ExcelWriter(out_path_found, engine="openpyxl") as writer:
                    df_found.to_excel(writer, index=False, sheet_name="WBS Preenchido")
                    _style_sheet(writer.sheets["WBS Preenchido"], kinds_found)
                log(f"\n✓ MQT_{ifc_stem}.xlsx exportado")

                out_path_full = out_dir / f"ElementosVerificados_{ifc_stem}.xlsx"
                n = 1
                while out_path_full.exists():
                    out_path_full = out_dir / f"ElementosVerificados_{ifc_stem}({n}).xlsx"
                    n += 1
                with pd.ExcelWriter(out_path_full, engine="openpyxl") as writer:
                    df_full.to_excel(writer, index=False, sheet_name="WBS Preenchido")
                    _style_sheet(writer.sheets["WBS Preenchido"], kinds_full)
                log(f"✓ WBS_ElementosMapeados_{ifc_stem}.xlsx exportado")

                out_path = out_path_full

                headers = [
                    "ifc_filename",
                    "wbs_codigo", "descricao",
                    "ifc_class", "predefinedtype", "objecttype",
                    "material",
                    "ifc_guid", "buildingstorey", "classification_code",
                    "ifc_project", "ifc_site", "ifc_building",
                    "ifc_valor", "unidade", "qty_type",
                ]

                wbs_rows = []
                df_w  = self.df_raw.copy()
                lvl_w = split_levels(df_w, self.col_nivel)
                for i in range(len(df_w)):
                    if lvl_w.iat[i] is not None and lvl_w.iat[i] < 10:
                        c_code = str(df_w.at[i, self.col_wbs]).strip() if pd.notna(df_w.at[i, self.col_wbs]) else ""
                        c_desc = str(df_w.at[i, self.col_desc]).strip() if pd.notna(df_w.at[i, self.col_desc]) else ""
                        if c_code:
                            wbs_rows.append({"wbs_codigo": c_code, "descricao": c_desc})

                per_code: dict[str, dict] = {}
                for det in all_details:
                    c = det.get("wbs_codigo")
                    if not c:
                        continue
                    blob = per_code.setdefault(c, {"meta": {}, "groups": {}})
                    blob["meta"] = {
                        "ifc_project":  det.get("ifc_project",  "n/a"),
                        "ifc_site":     det.get("ifc_site",     "n/a"),
                        "ifc_building": det.get("ifc_building", "n/a"),
                    }
                    gval = det.get("group_value") or "n/a"
                    blob["groups"].setdefault(gval, []).append({
                        "ifc_class":           det.get("ifc_class",           "n/a"),
                        "predefined":          det.get("predefinedtype",       "n/a"),
                        "objecttype":          det.get("objecttype",           "n/a"),
                        "material":            det.get("material",             "n/a"),
                        "guid":                det.get("ifc_guid",             "n/a"),
                        "buildingstorey":      det.get("buildingstorey",       "n/a"),
                        "classification_code": det.get("classification_code",  "n/a"),
                        "value":               det.get("ifc_valor",            ""),
                        "qty_type":            det.get("qty_type",             "prop"),
                    })

                for c in no_elements_codes:
                    per_code.setdefault(c, {"meta": {}, "groups": {}, "no_elements": True})
                    per_code[c]["no_elements"] = True

                self._last_csv_cache = {
                    "headers":       headers,
                    "wbs_rows":      wbs_rows,
                    "per_code":      per_code,
                    "code_to_unit":  code_to_unit,
                }
                self.last_detailed_rows = all_details
                self.page_report.btn_export_csv.config(state="normal")

                log(f"Detalhes recolhidos: {len(all_details)} elementos")
                if no_elements_codes:
                    log(f"⚠  Códigos sem elementos: {len(no_elements_codes)}")

                if on_finish:
                    finish_msg = "Dois ficheiros exportados com sucesso."
                    if no_elements_codes:
                        finish_msg += (
                            f"\n\nNão foram encontrados elementos equivalentes"
                            f" a {len(no_elements_codes)} itens do mapeamento carregado."
                        )
                    on_finish(finish_msg)

            except Exception as e:
                import traceback
                traceback.print_exc()
                if on_finish:
                    on_finish(f"Erro: {e}")

        threading.Thread(target=worker, daemon=True).start()
